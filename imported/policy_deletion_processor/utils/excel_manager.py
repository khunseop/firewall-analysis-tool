#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel 파일 관리 기능을 제공하는 모듈
"""

import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

logger = logging.getLogger(__name__)

class ExcelManager:
    """Excel 파일 관리 기능을 제공하는 클래스"""
    
    def __init__(self, config_manager):
        """
        Excel 관리자를 초기화합니다.
        
        Args:
            config_manager: 설정 관리자
        """
        self.config = config_manager
    
    def save_to_excel(self, df, sheet_type, file_name):
        """
        DataFrame을 Excel 파일에 저장합니다.
        
        Args:
            df: 저장할 DataFrame
            sheet_type (str): 시트 유형
            file_name (str): 파일 이름
        """
        try:
            wb = load_workbook(file_name)
            sheet = wb[sheet_type]
            
            # 첫 번째 행 삽입
            sheet.insert_rows(1)
            sheet['A1'] = '="대상 정책 수: "&COUNTA(B:B)-1'
            sheet['A1'].font = Font(bold=True)
            
            # 헤더 스타일 설정
            header_color = self.config.get('excel_styles.header_fill_color', 'E0E0E0')
            history_color = self.config.get('excel_styles.history_fill_color', 'ccffff')
            
            for col in range(1, 8):
                cell = sheet.cell(row=2, column=col)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            
            if sheet_type != '이력없음_미사용정책':
                for col in range(8, 24):
                    cell = sheet.cell(row=2, column=col)
                    cell.fill = PatternFill(start_color=history_color, end_color=history_color, fill_type='solid')
            
            wb.save(file_name)
            logger.info(f"Excel 파일 '{file_name}'의 '{sheet_type}' 시트에 데이터를 저장했습니다.")
        except Exception as e:
            logger.exception(f"Excel 파일 저장 중 오류 발생: {e}")
            raise 