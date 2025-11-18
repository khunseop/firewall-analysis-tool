import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font

# 컬럼 순서 재정렬 (구분, 객체 타입 앞에)
def reorder_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = df.columns.tolist()
    for col in ['객체 타입', '구분']:
        if col in cols:
            cols.insert(0, cols.pop(cols.index(col)))
    return df[cols]

def save_results_to_excel(
    added_df: pd.DataFrame,
    removed_df: pd.DataFrame,
    modified_list: list,
    object_diffs: dict,
    output_file: str,
    changed_obj_names: dict,
    df_old: pd.DataFrame
) -> str:
    # 정책 결과 분리
    added_removed_records = []
    modified_records = []

    for _, row in added_df.iterrows():
        row_dict = row.to_dict()
        row_dict['구분'] = '추가'
        added_removed_records.append(row_dict)

    for _, row in removed_df.iterrows():
        row_dict = row.to_dict()
        row_dict['구분'] = '삭제'
        added_removed_records.append(row_dict)

    for item in modified_list:
        rule = item['Rule Name']
        if item['Changes']:
            for field, change in item['Changes'].items():
                modified_records.append({
                    'Rule Name': rule,
                    'Field': field,
                    'From': ', '.join(change['from']) if isinstance(change['from'], list) else change['from'],
                    'To': ', '.join(change['to']) if isinstance(change['to'], list) else change['to'],
                    'Added Items': ', '.join(change['added']),
                    'Removed Items': ', '.join(change['removed']),
                    'Indirect Change': item['Indirect Change'],
                    '구분': '변경'
                })
        else:
            for field, refs in item['Indirect Fields']:
                affected_str = ', '.join(sorted(refs))
                modified_records.append({
                    'Rule Name': rule,
                    'Field': field,
                    'From': affected_str,
                    'To': affected_str,
                    'Added Items': '',
                    'Removed Items': '',
                    'Indirect Change': True,
                    '구분': '변경'
                })

    policy_added_removed_df = pd.DataFrame(added_removed_records)
    policy_modified_df = pd.DataFrame(modified_records)

    policy_added_removed_df = reorder_columns(policy_added_removed_df)
    policy_modified_df = reorder_columns(policy_modified_df)

    # 객체 결과 분리
    object_added_removed_records = []
    object_modified_records = []
    for obj_type, (added, removed, modified) in object_diffs.items():
        cleaned_type = obj_type.replace('_diff', '')

        for item in added:
            record = item.copy()
            record['구분'] = '추가'
            record['객체 타입'] = cleaned_type
            object_added_removed_records.append(record)
        for item in removed:
            record = item.copy()
            record['구분'] = '삭제'
            record['객체 타입'] = cleaned_type
            object_added_removed_records.append(record)
        for item in modified:
            record = item.copy()
            record['구분'] = '변경'
            record['객체 타입'] = cleaned_type
            object_modified_records.append(record)

    object_added_removed_df = pd.DataFrame(object_added_removed_records)
    object_modified_df = pd.DataFrame(object_modified_records)

    # 컬럼 순서: 구분, 객체 타입을 앞에
    object_added_removed_df = reorder_columns(object_added_removed_df)
    object_modified_df = reorder_columns(object_modified_df)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        policy_added_removed_df.to_excel(writer, sheet_name='정책 증감', index=False)
        policy_modified_df.to_excel(writer, sheet_name='정책 변경', index=False)
        object_added_removed_df.to_excel(writer, sheet_name='객체 증감', index=False)
        object_modified_df.to_excel(writer, sheet_name='객체 변경', index=False)

    _adjust_excel_formatting(output_file)
    return output_file


def _adjust_excel_formatting(excel_path: str, max_width: int = 80) -> None:
    wb = load_workbook(excel_path)

    summary = wb.create_sheet(title="Summary", index=0)
    summary.append(["시트명", "행 개수", "설명"])

    # 헤더 스타일: 회색 + bold
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    DESC_MAP = {
        "정책 증감": "추가되거나 삭제된 정책 항목 목록",
        "정책 변경": "필드 값이 수정된 정책 항목 목록",
        "객체 증감": "Address/Service 등 객체의 추가 또는 삭제된 항목",
        "객체 변경": "객체 속성 변경 또는 구성 변경 내역"
    }

    for ws in wb.worksheets[1:]:
        desc = DESC_MAP.get(ws.title, "정책/객체 비교 결과")
        summary.append([ws.title, ws.max_row - 1, desc])

    color_map = {
        '추가': "C6EFCE",
        '삭제': "FFC7CE",
        '변경': "FFEB9C"
    }

    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, str) and cell.value in color_map:
                    cell.fill = PatternFill(start_color=color_map[cell.value], fill_type="solid")

    wb.save(excel_path)
