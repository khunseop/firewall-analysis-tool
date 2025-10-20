import os
import re
import logging
import paramiko
from scp import SCPClient
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Paramiko의 로그 레벨을 WARNING 이상으로 설정 (INFO 로그 제거)
logging.getLogger("paramiko").setLevel(logging.WARNING)

# 명령어 상수
POLICY_DIRECTORY = 'ls -ls *.fwrules'
CONF_DIRECTORY = 'ls *.conf'
INFO_FILE = 'cat /etc/SECUIMF2.info'

# 정규표현식 패턴
HOST_PATTERN = {
    'id': r'id = (\d+)',
    'name': r'name = "([^"]+)"',
    'zone': r'zone = "([^"]+)"',
    'user': r'user = "([^"]+)"',
    'date': r'date = "([^"]+)"',
    'ip': r'ip = "([^"]+)"',
    'description': r'd = "([^"]+)"',
}
MASK_PATTERN = {
    'id': r'id = (\d+)',
    'name': r'name = "([^"]+)"',
    'zone': r'zone = "([^"]+)"',
    'user': r'user = "([^"]+)"',
    'date': r'date = "([^"]+)"',
    'ip/start': r'ip="([^"]+)"',
    'mask/end': r'mask="([^"]+)"',
    'description': r'd = "([^"]+)"',
}
RANGE_PATTERN = {
    'id': r'id = (\d+)',
    'name': r'name = "([^"]+)"',
    'zone': r'zone = "([^"]+)"',
    'user': r'user = "([^"]+)"',
    'date': r'date = "([^"]+)"',
    'ip/start': r'rangestart="([^"]+)"',
    'mask/end': r'rangeend="([^"]+)"',
    'description': r'd = "([^"]+)"',
}
GROUP_PATTERN = {
    'id': r'id = (\d+)',
    'name': r'name = "([^"]+)"',
    'zone': r'zone = "([^"]+)"',
    'user': r'user = "([^"]+)"',
    'date': r'date = "([^"]+)"',
    'count': r'count = \{(.*?)\},',
    'hosts': r'hosts=\{(.*?)\},',
    'networks': r'networks=\{(.*?)\},',
    'description': r'd = "([^"]+)"',
}
SERVICE_PATTERN = {
    'id': r'id = (\d+)',
    'name': r'name = "([^"]+)"',
    'protocol': r'protocol="([^"]+)",',
    'str_src_port': r'str_src_port="([^"]+)",',
    'str_svc_port': r'str_svc_port="([^"]+)",',
    'svc_type': r'svc_type="([^"]+)",',
    'description': r'd = "([^"]+)"',
}

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')


# ────────────── HELPER FUNCTIONS ──────────────

def create_ssh_client(host: str, port: int, username: str, password: str) -> paramiko.SSHClient:
    """
    SSHClient를 생성하고 연결한 후 반환합니다.
    """
    client = paramiko.SSHClient()
    client.load_system_host_keys()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(host, port, username, password)
    return client


def exec_remote_command(ssh: paramiko.SSHClient, command: str, remote_directory: str = None):
    """
    원격 디렉토리 변경 후 명령어 실행
    """
    full_command = f'cd {remote_directory} && {command}' if remote_directory else command
    return ssh.exec_command(full_command)


def download_file(ssh: paramiko.SSHClient, remote_directory: str, file_name: str, local_directory: str, host: str) -> str:
    """
    SCPClient를 사용하여 파일을 다운로드하고, 다운로드된 파일명을 반환합니다.
    """
    remote_path = os.path.join(remote_directory, file_name)
    download_name = f"{host}_{file_name}"
    local_path = os.path.join(local_directory, download_name)
    with SCPClient(ssh.get_transport()) as scp:
        scp.get(remote_path, local_path)
    return download_name


# ────────────── SSH/FILE DOWNLOAD FUNCTIONS ──────────────

def export_mf2_data(host: str, port: int, username: str, password: str,
                    remote_directory: str, local_directory: str) -> list:
    downloaded_files = []
    ssh = None
    try:
        ssh = create_ssh_client(host, port, username, password)
        # fwrules 파일 다운로드
        _, stdout, stderr = exec_remote_command(ssh, POLICY_DIRECTORY, remote_directory)
        if stderr.read():
            raise Exception(f"정책 파일 조회 실패: {stderr.read().decode()}")
            
        fwrules_lines = stdout.readlines()
        if fwrules_lines:
            latest_file = fwrules_lines[0].split()[-1]
            downloaded_files.append(download_file(ssh, remote_directory, latest_file, local_directory, host))

        # conf 파일 다운로드
        specified_conf_files = ['groupobject.conf', 'hostobject.conf', 'networkobject.conf', 'serviceobject.conf']
        _, stdout, stderr = exec_remote_command(ssh, CONF_DIRECTORY, remote_directory)
        if stderr.read():
            raise Exception(f"설정 파일 조회 실패: {stderr.read().decode()}")
            
        conf_lines = stdout.readlines()
        for line in conf_lines:
            conf_file = line.strip()
            if conf_file in specified_conf_files:
                downloaded_files.append(download_file(ssh, remote_directory, conf_file, local_directory, host))
        
        return downloaded_files
    except Exception as e:
        logging.error(f"MF2 데이터 내보내기 실패 - {host}: {str(e)}")
        raise Exception(f"MF2 데이터 내보내기 실패: {str(e)}")
    finally:
        if ssh:
            ssh.close()


def download_rule_file(host: str, port: int, username: str, password: str,
                       remote_directory: str, local_directory: str) -> str:
    """
    원격 장비에서 fwrules 파일(최신 파일 1건)을 다운로드한 후,
    다운로드된 파일명을 반환합니다.
    """
    latest_download = ""
    ssh = create_ssh_client(host, port, username, password)
    try:
        _, stdout, _ = exec_remote_command(ssh, POLICY_DIRECTORY, remote_directory)
        fwrules_lines = stdout.readlines()
        if fwrules_lines:
            latest_file = fwrules_lines[0].split()[-1]
            latest_download = download_file(ssh, remote_directory, latest_file, local_directory, host)
    except Exception as e:
        logging.error("download_rule_file error: %s", e)
    finally:
        ssh.close()
        return latest_download


def download_object_files(host: str, port: int, username: str, password: str,
                          remote_directory: str, local_directory: str, conf_types: list = None) -> list:
    """
    원격 장비에서 지정된 conf 파일들을 다운로드한 후, 다운로드된 파일명 리스트를 반환합니다.
    """
    if conf_types is None:
        conf_types = ['groupobject.conf', 'hostobject.conf', 'networkobject.conf', 'serviceobject.conf']
    
    downloaded_files = []
    ssh = create_ssh_client(host, port, username, password)
    try:
        _, stdout, _ = exec_remote_command(ssh, CONF_DIRECTORY, remote_directory)
        conf_lines = stdout.readlines()
        with SCPClient(ssh.get_transport()) as scp:
            for line in conf_lines:
                conf_file = line.strip()
                if conf_file in conf_types:
                    # 다운로드할 로컬 파일 경로 지정
                    download_name = f"{host}_{conf_file}"
                    local_path = os.path.join(local_directory, download_name)
                    # 이미 파일이 있으면 다운로드하지 않음
                    if not os.path.exists(local_path):
                        scp.get(os.path.join(remote_directory, conf_file), local_path)
                    downloaded_files.append(local_path)
    except Exception as e:
        logging.error("download_object_files error: %s", e)
    finally:
        ssh.close()
        return downloaded_files


def show_system_info(host: str, username: str, password: str) -> pd.DataFrame:
    """
    원격 장비의 시스템 정보를 수집하여 DataFrame으로 반환합니다.
    """
    ssh = create_ssh_client(host, 22, username, password)
    try:
        # hostname
        _, stdout, _ = ssh.exec_command('hostname')
        hostname = stdout.readline().strip()

        # uptime (공백 기준 분할 후 4번째, 5번째 요소 사용)
        _, stdout, _ = ssh.exec_command('uptime')
        uptime_parts = stdout.readline().rstrip().split(' ')
        uptime = f"{uptime_parts[3]} {uptime_parts[4].rstrip(',')}" if len(uptime_parts) >= 5 else ""

        # SECUIMF2 정보
        _, stdout, _ = ssh.exec_command(INFO_FILE)
        info_lines = stdout.readlines()
        # rpm version
        _, stdout, _ = ssh.exec_command('rpm -q mf2')
        version = stdout.readline().strip()

        # info_lines 순서에 따라 모델, mac, serial 추출
        model = info_lines[0].split('=')[1].strip() if len(info_lines) > 0 else ""
        mac_address = info_lines[2].split('=')[1].strip() if len(info_lines) > 2 else ""
        hw_serial = info_lines[3].split('=')[1].strip() if len(info_lines) > 3 else ""

        data = {
            "hostname": hostname,
            "ip_address": host,
            "mac_address": mac_address,
            "uptime": uptime,
            "model": model,
            "serial_number": hw_serial,
            "sw_version": version,
        }
        return pd.DataFrame(data, index=[0])
    except Exception as e:
        logging.error("show_system_info error: %s", e)
    finally:
        ssh.close()


def delete_files(file_paths):
    """
    파일 경로(리스트 또는 단일 경로)에 대해 존재하면 삭제합니다.
    """
    if not isinstance(file_paths, list):
        file_paths = [file_paths]
    for path in file_paths:
        if os.path.exists(path):
            try:
                os.remove(path)
            except Exception as e:
                logging.error("파일 삭제 실패 (%s): %s", path, e)
        else:
            logging.warning("File not found: %s", path)


# ────────────── FILE CONTENT & PARSING FUNCTIONS ──────────────

def remove_newlines_from_file(file_path: str) -> str:
    """
    파일 내용을 읽어와 모든 개행문자를 제거한 문자열을 반환합니다.
    """
    try:
        with open(file_path, 'r', encoding='utf-8-sig') as file:
            content = file.read()
        return content.replace('\n', '')
    except Exception as e:
        logging.error("remove_newlines_from_file error: %s", e)
        return f"error: {e}"


def extract_braces_of_depth_1_or_more(content: str) -> list:
    """
    중괄호({})로 둘러싸인 블록 중 깊이가 1 이상인 내용들을 리스트로 반환합니다.
    """
    depth = 0
    results = []
    temp = ""
    for char in content:
        if char == '{':
            if depth == 0:
                temp = ""
            temp += char
            depth += 1
        elif char == '}':
            temp += char
            depth -= 1
            if depth == 0:
                results.append(temp.strip())
        elif depth >= 1:
            temp += char
    return results


def extract_braces_of_depth_2_or_more_without_outer_braces(content: str) -> list:
    """
    중괄호 블록 중 깊이가 2 이상인 부분만 추출하여 외부 중괄호는 제거한 내용을 리스트로 반환합니다.
    """
    depth = 0
    results = []
    temp = ""
    for char in content:
        if char == '{':
            if depth >= 1:
                temp += char
            depth += 1
        elif char == '}':
            depth -= 1
            if depth >= 1:
                temp += char
                if depth == 1:
                    results.append(temp[1:-1].strip())
                    temp = ""
        elif depth >= 2:
            temp += char
    return results


def parse_object(input_str: str) -> str:
    """
    문자열에서 따옴표를 제거한 후, 공백이나 콤마 기준으로 두 번째 토큰을 추출하여 콤마로 연결한 문자열을 반환합니다.
    """
    cleaned = input_str.replace('"', '')
    parsed = []
    if "," in cleaned:
        for entry in cleaned.split(','):
            parts = entry.split(' ')
            if len(parts) > 1:
                parsed.append(parts[1])
    elif " " in cleaned:
        parts = cleaned.split(' ')
        if len(parts) > 1:
            parsed.append(parts[1])
    else:
        parsed.append(cleaned)
    return ','.join(parsed)


def group_parsing(file_path: str) -> pd.DataFrame:
    """
    그룹 객체 파일을 파싱하여 DataFrame으로 반환합니다.
    """
    content = remove_newlines_from_file(file_path)
    depth_braces = extract_braces_of_depth_2_or_more_without_outer_braces(content)
    if depth_braces:
        depth_braces.pop(0)  # id 정보 삭제

    data_list = []
    for text in depth_braces:
        data = {}
        for key, pattern in GROUP_PATTERN.items():
            match = re.search(pattern, text)
            if match:
                if key in ['hosts', 'networks']:
                    items = []
                    obj_str = match.group(1)
                    if obj_str:
                        for item in obj_str.split(','):
                            # item 형식: key=value 또는 [key]
                            items.append(item.split('=')[0].replace('[', '').replace(']', ''))
                    data[key] = ','.join(items)
                elif key == 'count':
                    items = []
                    obj_str = match.group(1)
                    if obj_str:
                        for item in obj_str.split(','):
                            parts = item.split('=')
                            if len(parts) > 1:
                                items.append(parts[1])
                    data[key] = ','.join(items)
                else:
                    data[key] = match.group(1)
        data_list.append(data)
    return pd.DataFrame(data_list)


def service_parsing(file_path: str) -> pd.DataFrame:
    """
    서비스 객체 파일을 파싱하여 DataFrame으로 반환합니다.
    """
    content = remove_newlines_from_file(file_path)
    depth_braces = extract_braces_of_depth_2_or_more_without_outer_braces(content)
    if depth_braces:
        # 첫 두 항목(id 등) 삭제
        depth_braces.pop(0)
    if depth_braces:
        depth_braces.pop(0)

    data_list = []
    for text in depth_braces:
        data = {}
        for key, pattern in SERVICE_PATTERN.items():
            match = re.search(pattern, text)
            if match:
                data[key] = match.group(1)
        data_list.append(data)
    return pd.DataFrame(data_list)


def network_parsing(file_path: str) -> pd.DataFrame:
    """
    네트워크 객체 파일을 파싱하여 DataFrame으로 반환합니다.
    range 문자열 포함 여부에 따라 RANGE_PATTERN 또는 MASK_PATTERN을 사용합니다.
    """
    content = remove_newlines_from_file(file_path)
    depth_braces = extract_braces_of_depth_2_or_more_without_outer_braces(content)
    if depth_braces:
        depth_braces.pop(0)
    data_list = []
    for text in depth_braces:
        data = {}
        pattern = RANGE_PATTERN if "range" in text else MASK_PATTERN
        for key, pat in pattern.items():
            match = re.search(pat, text)
            if match:
                data[key] = match.group(1)
        data_list.append(data)
    return pd.DataFrame(data_list)


def host_parsing(file_path: str) -> pd.DataFrame:
    """
    호스트 객체 파일을 파싱하여 DataFrame으로 반환합니다.
    """
    content = remove_newlines_from_file(file_path)
    depth_braces = extract_braces_of_depth_2_or_more_without_outer_braces(content)
    if depth_braces:
        depth_braces.pop(0)
    data_list = []
    for text in depth_braces:
        data = {}
        for key, pattern in HOST_PATTERN.items():
            match = re.search(pattern, text)
            if match:
                data[key] = match.group(1)
        data_list.append(data)
    return pd.DataFrame(data_list)


def rule_parsing(file_path: str) -> pd.DataFrame:
    """
    규칙(rule) 파일을 파싱하여 DataFrame으로 반환합니다.
    """
    content = remove_newlines_from_file(file_path)
    depth_braces = extract_braces_of_depth_2_or_more_without_outer_braces(content)
    if not depth_braces:
        return pd.DataFrame()
    rule_blocks = extract_braces_of_depth_1_or_more(depth_braces[0])

    # 정규표현식 패턴들
    rule_pattern = r"\{rid=(.*?), "
    description_pattern = r"description=\"(.*?)\", use="
    use_pattern = r"use=\"(.*?)\", action"
    action_pattern = r"action=\"(.*?)\", group"
    shaping_pattern = r"shaping_string=\"(.*?)\", bi_di"
    source_pattern = r"from = \{(.*?)\},  to"
    destination_pattern = r"to = \{(.*?)\},  service"
    service_pattern = r"service = \{(.*?)\},  vid"
    ua_pattern = r"ua = \{(.*?)\}, unuse"

    policies = []
    for idx, block in enumerate(rule_blocks):
        rulename = re.findall(rule_pattern, block)
        description = re.findall(description_pattern, block)
        use = re.findall(use_pattern, block)
        action = re.findall(action_pattern, block)
        shaping_string = re.findall(shaping_pattern, block)
        shaping_string = shaping_string[0] if shaping_string else ""
        schedule = shaping_string.split('=')[1].lstrip('"') if "time=" in shaping_string else ''
        source = re.findall(source_pattern, block)
        destination = re.findall(destination_pattern, block)
        service = re.findall(service_pattern, block)
        ua = re.findall(ua_pattern, block)

        policy = {
            "Seq": idx + 1,
            "Rule Name": int(rulename[0]) if rulename else None,
            "Enable": use[0] if use else "",
            "Action": action[0] if action else "",
            "Source": parse_object(source[0]) if source else "",
            "User": parse_object(ua[0]) if ua else "",
            "Destination": parse_object(destination[0]) if destination else "",
            "Service": parse_object(service[0]) if service else "",
            "Application": "Any",
            "Security Profile": schedule,
            "Description": description[0] if description else "",
        }
        policies.append(policy)

    df = pd.DataFrame(policies)
    # 빈 문자열 또는 공백은 "Any"로 치환
    for col in ['Source', 'Destination', 'Service', 'User']:
        df[col] = df[col].replace({'': 'Any', ' ': 'Any'})
    return df


# ────────────── OBJECT EXPORT & COMBINE FUNCTIONS ──────────────

def combine_mask_end(row: pd.Series) -> str:
    """
    네트워크 객체에서 ip/start와 mask/end 값을 결합합니다.
    mask/end가 숫자면 cidr 표기, 아니면 범위 표기합니다.
    """
    if row.get('mask/end', '').isdigit():
        return f"{row.get('ip/start')}/{row.get('mask/end')}"
    else:
        return f"{row.get('ip/start')}-{row.get('mask/end')}"


def replace_values(ids: str, mapping: dict) -> str:
    """
    콤마로 구분된 id 문자열을 mapping 사전을 통해 값으로 치환하여 반환합니다.
    """
    return ','.join(mapping.get(item.strip(), '') for item in ids.split(','))


def combine_group_objects(row: pd.Series) -> str:
    """
    그룹 객체에서 변환된 hosts와 networks 값을 결합합니다.
    """
    values = [row.get('convert_hosts', ''), row.get('convert_networks', '')]
    return ','.join(val for val in values if val and val.strip())


def export_address_objects(group_file: str, host_file: str, network_file: str) -> tuple:
    """
    그룹, 호스트, 네트워크 객체 파일을 파싱하여
    네트워크 객체(DataFrame)와 그룹 객체(DataFrame)를 반환합니다.
    """
    group_df = group_parsing(group_file)
    network_df = network_parsing(network_file)
    host_df = host_parsing(host_file)

    if not network_df.empty:
        network_df['Value'] = network_df.apply(combine_mask_end, axis=1)
    network_ids = dict(zip(network_df['id'].astype(str), network_df['Value'])) if 'id' in network_df and 'Value' in network_df else {}
    host_ids = dict(zip(host_df['id'].astype(str), host_df['ip'])) if 'id' in host_df and 'ip' in host_df else {}

    group_df['convert_networks'] = group_df['networks'].apply(lambda x: replace_values(x, network_ids)) if 'networks' in group_df else ""
    group_df['convert_hosts'] = group_df['hosts'].apply(lambda x: replace_values(x, host_ids)) if 'hosts' in group_df else ""
    group_df['Entry'] = group_df.apply(combine_group_objects, axis=1)

    # 필요한 컬럼 선택 및 이름 변경
    group_df = group_df[['name', 'Entry']]
    group_df.columns = ['Group Name', 'Entry']
    network_df = network_df[['name', 'Value']]
    network_df.columns = ['Name', 'Value']
    host_df = host_df[['name', 'ip']]
    host_df.columns = ['Name', 'Value']
    network_objects_df = pd.concat([host_df, network_df], axis=0, ignore_index=True)

    return network_objects_df, group_df


def export_service_objects(service_file: str) -> pd.DataFrame:
    """
    서비스 객체 파일을 파싱하여 DataFrame으로 반환합니다.
    """
    service_df = service_parsing(service_file)
    if not service_df.empty:
        service_df = service_df[['name', 'protocol', 'str_svc_port']]
        service_df.columns = ['Name', 'Protocol', 'Port']
    return service_df


def export_objects(device_ip: str, username: str, password: str) -> list:
    """
    원격 장비에서 객체 파일(conf)들을 다운로드하여 그룹/호스트/네트워크, 서비스 DataFrame을 생성한 후,
    다운로드된 파일들은 삭제하고 DataFrame 리스트를 반환합니다.
    """
    files = download_object_files(device_ip, 22, username, password, '/secui/etc/', './')
    if len(files) < 4:
        logging.error("필요한 conf 파일이 모두 다운로드되지 않았습니다.")
        return []
    group_file, host_file, network_file, service_file = files[:4]
    address_df, address_group_df = export_address_objects(group_file, host_file, network_file)
    service_df = export_service_objects(service_file)
    delete_files(files)
    return [address_df, address_group_df, service_df]


def export_security_rules(device_ip: str, username: str, password: str) -> pd.DataFrame:
    """
    원격 장비에서 규칙 파일(fwrules)을 다운로드하여 파싱한 후 DataFrame으로 반환합니다.
    """
    file_name = download_rule_file(device_ip, 22, username, password, '/secui/etc/', './')
    if not file_name:
        logging.error("규칙 파일 다운로드 실패")
        return pd.DataFrame()
    rule_df = rule_parsing(file_name)
    delete_files(file_name)
    return rule_df


# ────────────── SAVE TO EXCEL FUNCTION ──────────────
def apply_excel_style(file_name: str) -> None:
    """
    모든 시트에 대해 헤더에 연한 회색 배경을 적용하고,
    헤더의 너비를 자동 조절하되 최대 너비를 40으로 제한합니다.
    
    :param file_name: 처리할 엑셀 파일 이름
    """
    try:
        workbook = load_workbook(file_name)
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        # 모든 워크시트를 순회하며 스타일 적용
        for worksheet in workbook.worksheets:
            header_row = worksheet[1]
            for cell in header_row:
                cell.fill = header_fill
                column_letter = cell.column_letter
                max_length = 0
                for col_cell in worksheet[column_letter]:
                    if col_cell.value is not None:
                        cell_length = len(str(col_cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) * 1.2
                worksheet.column_dimensions[column_letter].width = min(40, adjusted_width)

        workbook.save(file_name)
    except Exception as error:
        logging.error("엑셀 스타일 적용 중 오류 발생: %s", error)

def save_dfs_to_excel(dfs, sheet_names, file_name: str) -> bool:
    """
    단일 DataFrame 또는 DataFrame 리스트를 지정된 시트명으로 엑셀 파일에 저장합니다.
    """
    try:
        if not isinstance(dfs, list):
            dfs = [dfs]
        if not isinstance(sheet_names, list):
            sheet_names = [sheet_names]
        with pd.ExcelWriter(file_name) as writer:
            for df, sheet in zip(dfs, sheet_names):
                df.to_excel(writer, sheet_name=sheet, index=False)
        apply_excel_style(file_name)
        return True
    except Exception as e:
        logging.error("save_dfs_to_excel error: %s", e)
        return False