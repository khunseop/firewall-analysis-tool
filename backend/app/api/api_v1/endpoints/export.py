import logging
import pandas as pd
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO
from urllib.parse import quote
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

router = APIRouter()

HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _hex(color: str) -> str:
    return color.lstrip("#").upper()


def _write_sheet(ws, columns: list, rows: list) -> None:
    """한 시트에 헤더 + 데이터 행을 채운다. _build_styled_excel/_build_multi_sheet_excel이 공유."""
    ws.append([col["header"] for col in columns])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
    ws.row_dimensions[1].height = 22

    for i, col in enumerate(columns):
        ws.column_dimensions[get_column_letter(i + 1)].width = col.get("width", 15)

    for row_data in rows:
        values = row_data.get("values", [])
        row_bg = row_data.get("rowBg")
        cell_colors = row_data.get("cellFontColors", [])

        ws.append(values)
        row_idx = ws.max_row
        ws.row_dimensions[row_idx].height = 18

        bg_fill = PatternFill(start_color=_hex(row_bg), end_color=_hex(row_bg), fill_type="solid") if row_bg else None

        for col_idx, cell in enumerate(ws[row_idx]):
            cell.alignment = Alignment(vertical="center")
            if bg_fill:
                cell.fill = bg_fill
            font_color = cell_colors[col_idx] if col_idx < len(cell_colors) else None
            if font_color:
                cell.font = Font(color=_hex(font_color), bold=True)


def _build_styled_excel(data: dict) -> BytesIO:
    """New structured format: {columns: [{header, width}], rows: [{values, rowBg, cellFontColors}]}"""
    wb = Workbook()
    ws = wb.active
    ws.title = "분석결과"
    _write_sheet(ws, data["columns"], data["rows"])

    output = BytesIO()
    wb.save(output)
    return output


def _build_multi_sheet_excel(data: dict) -> BytesIO:
    """다중 시트 포맷: {sheets: [{name, columns, rows}, ...]}"""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in data["sheets"]:
        # 엑셀 시트명은 31자 제한 + 일부 특수문자 금지라 안전하게 자른다.
        ws = wb.create_sheet(title=sheet["name"][:31])
        _write_sheet(ws, sheet["columns"], sheet["rows"])

    output = BytesIO()
    wb.save(output)
    return output


def _build_flat_excel(data: dict) -> BytesIO:
    """Original flat format (backwards compat for Policies / Objects pages)."""
    rows = data.get("data", [])
    if not rows:
        raise HTTPException(status_code=400, detail="No data to export")

    df = pd.DataFrame(rows)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]

        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
        ws.row_dimensions[1].height = 20

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    return output


@router.post("/export/excel")
async def export_to_excel(data: dict):
    try:
        filename = data.get("filename", "export")
        if "sheets" in data:
            output = _build_multi_sheet_excel(data)
        elif "columns" in data and "rows" in data:
            output = _build_styled_excel(data)
        else:
            output = _build_flat_excel(data)

        output.seek(0)
        encoded_name = quote(f"{filename}.xlsx", safe="")
        headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"}
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except HTTPException:
        raise
    except Exception as e:
        logging.error(f"Failed to export Excel: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")
