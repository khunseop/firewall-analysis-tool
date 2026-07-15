import datetime
import json
import logging
import os
from io import BytesIO
from typing import Any, Dict, List
from urllib.parse import quote

import openpyxl
import yaml
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel
from sqlalchemy.ext.asyncio import AsyncSession

from app import crud, schemas
from app.db.session import get_db

logger = logging.getLogger(__name__)
router = APIRouter()

# fpat.yaml 경로
_FPAT_YAML = os.path.abspath(os.path.join(
    os.path.dirname(__file__),
    '..', '..', '..', '..', '..',
    'fpat', 'fpat.yaml',
))

_SETTINGS_KEY = 'deletion_workflow_config'


# ──────────────────────────────────────────────
# 헬퍼 함수
# ──────────────────────────────────────────────

def _default_config() -> dict:
    """UI 렌더링용 기본 설정 (fpat.yaml 구조와 동일)"""
    return {
        # ── 파일 관리 ───────────────────────────────────────────────
        "file_management": {
            "policy_version_format": "_v{version}",
            "final_version_suffix": "_vf",
            "default_extension": ".xlsx",
        },
        # request_extractor.py: config.get('file_naming.request_id_prefix')
        "file_naming": {
            "request_id_prefix": "GSAMS신청번호_",
        },
        # mis_id_adder.py: config.get('file_extensions.csv')
        "file_extensions": {
            "csv": ".csv",
        },
        # ── 분석 기준 ───────────────────────────────────────────────
        "analysis_criteria": {
            "unused_threshold_days": 90,
        },
        # exception_handler.py: config.get('timeframes.recent_policy_days')
        "timeframes": {
            "recent_policy_days": 90,
        },
        # ── 예외 ────────────────────────────────────────────────────
        "exceptions": {
            "request_ids": [],
            "policy_rules": [],
            "static_list": [],
            "duplicate_policies": [],
        },
        # ── 정책 처리 ───────────────────────────────────────────────
        "policy_processing": {
            "request_parsing": {
                "gsams_3_pattern": "",
                "gsams_1_rulename_pattern": "",
                "gsams_1_user_pattern": "",
                "gsams_1_desc_pattern": "",
                "gsams_1_date_pattern": "",
            },
            "analysis_markers": {
                "paloalto": {
                    "deny_standard_rule_name": "",
                    "infrastructure_prefixes": [],
                    "infrastructure_exception_label": "인프라정책",
                    "special_policy_label": "라인그룹정책",
                },
                "secui": {
                    "deny_standard_description_keyword": "",
                    "infrastructure_exception_label": "인프라정책",
                },
            },
            "aggregation": {
                "column_mapping": {},
                "final_columns": [],
                "email_domain_map": {},
                "title_bracket_pattern": "",
            },
        },
        # ── 엑셀 스타일 ─────────────────────────────────────────────
        "excel_styles": {
            "header_fill_color": "E0E0E0",
            "history_fill_color": "CCFFFF",
        },
    }


def _deep_merge(base: dict, override: dict) -> dict:
    """base에 override를 재귀적으로 병합합니다. override 값이 우선합니다."""
    result = dict(base)
    for k, v in override.items():
        if k in result and isinstance(result[k], dict) and isinstance(v, dict):
            result[k] = _deep_merge(result[k], v)
        else:
            result[k] = v
    return result


def _load_fpat_yaml() -> dict:
    """fpat.yaml 파일에서 설정을 로드합니다. 없으면 기본값 반환."""
    if not os.path.exists(_FPAT_YAML):
        return _default_config()
    try:
        with open(_FPAT_YAML, 'r', encoding='utf-8') as f:
            data = yaml.safe_load(f) or {}
        return _deep_merge(_default_config(), data)
    except Exception as e:
        logger.warning(f"fpat.yaml 로드 실패: {e}")
        return _default_config()


def _write_fpat_yaml(config: dict) -> None:
    """설정을 fpat.yaml 파일에 동기화합니다."""
    if not os.path.exists(os.path.dirname(_FPAT_YAML)):
        return
    try:
        with open(_FPAT_YAML, 'w', encoding='utf-8') as f:
            yaml.dump(
                config, f,
                allow_unicode=True,
                default_flow_style=False,
                sort_keys=False,
            )
        logger.info(f"fpat.yaml 업데이트: {_FPAT_YAML}")
    except Exception as e:
        logger.warning(f"fpat.yaml 쓰기 실패 (DB는 저장됨): {e}")


# ──────────────────────────────────────────────
# 일반 설정 엔드포인트
# ──────────────────────────────────────────────

@router.get("/", response_model=List[schemas.Settings])
async def read_settings(db: AsyncSession = Depends(get_db)):
    """모든 설정 조회"""
    return await crud.settings.get_all_settings(db)


@router.get("/{key}", response_model=schemas.Settings)
async def read_setting(key: str, db: AsyncSession = Depends(get_db)):
    """특정 설정 조회"""
    setting = await crud.settings.get_setting(db, key=key)
    if setting is None:
        raise HTTPException(status_code=404, detail="Setting not found")
    return setting


@router.put("/{key}", response_model=schemas.Settings)
async def update_setting(
    key: str,
    setting_in: schemas.SettingsUpdate,
    db: AsyncSession = Depends(get_db),
):
    """설정 업데이트 (없으면 생성)"""
    setting = await crud.settings.get_setting(db, key=key)
    if setting is None:
        setting_create = schemas.SettingsCreate(
            key=key,
            value=setting_in.value,
            description=setting_in.description,
        )
        created_setting = await crud.settings.create_setting(db, setting_create)
        if key == "sync_parallel_limit":
            from app.services.sync.tasks import reset_sync_semaphore
            await reset_sync_semaphore()
        return created_setting

    updated_setting = await crud.settings.update_setting(
        db=db, db_obj=setting, obj_in=setting_in
    )
    if key == "sync_parallel_limit":
        from app.services.sync.tasks import reset_sync_semaphore
        await reset_sync_semaphore()
    return updated_setting


# ──────────────────────────────────────────────
# 정책 삭제 워크플로우 설정 엔드포인트
# ──────────────────────────────────────────────

class DeletionWorkflowConfigPayload(BaseModel):
    config: Dict[str, Any]


@router.get("/deletion-workflow/config")
async def get_deletion_workflow_config(db: AsyncSession = Depends(get_db)):
    """
    정책 삭제 워크플로우 설정 조회 (fpat.yaml 구조).

    우선순위: DB → fpat.yaml 파일 → 기본값
    """
    setting = await crud.settings.get_setting(db, key=_SETTINGS_KEY)
    if setting:
        try:
            stored = json.loads(setting.value)
            return _deep_merge(_default_config(), stored)
        except Exception:
            pass
    return _load_fpat_yaml()


@router.put("/deletion-workflow/config")
async def update_deletion_workflow_config(
    payload: DeletionWorkflowConfigPayload,
    db: AsyncSession = Depends(get_db),
):
    """
    정책 삭제 워크플로우 설정 저장 (fpat.yaml 구조).

    DB에 저장하고 fpat.yaml 파일에도 동기화합니다.
    """
    value = json.dumps(payload.config, ensure_ascii=False)
    setting = await crud.settings.get_setting(db, key=_SETTINGS_KEY)
    if setting is None:
        await crud.settings.create_setting(
            db,
            schemas.SettingsCreate(
                key=_SETTINGS_KEY,
                value=value,
                description='정책 삭제 워크플로우 설정 (fpat.yaml 형식)',
            ),
        )
    else:
        await crud.settings.update_setting(
            db=db,
            db_obj=setting,
            obj_in=schemas.SettingsUpdate(value=value),
        )

    _write_fpat_yaml(payload.config)
    return payload.config


@router.get("/deletion-workflow/config/export")
async def export_deletion_workflow_config(db: AsyncSession = Depends(get_db)):
    """현재 삭제 워크플로우 설정을 JSON 파일로 다운로드합니다."""
    setting = await crud.settings.get_setting(db, key=_SETTINGS_KEY)
    if setting:
        try:
            config = _deep_merge(_default_config(), json.loads(setting.value))
        except Exception:
            config = _load_fpat_yaml()
    else:
        config = _load_fpat_yaml()

    today = datetime.date.today().strftime("%Y%m%d")
    filename = f"deletion_workflow_config_{today}.json"
    content = json.dumps({"version": 1, "exported_at": today, "config": config},
                         ensure_ascii=False, indent=2).encode("utf-8")
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename, safe='')}"},
    )


@router.post("/deletion-workflow/config/import")
async def import_deletion_workflow_config(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """JSON 백업 파일에서 삭제 워크플로우 설정을 복구합니다."""
    raw = await file.read()
    try:
        data = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=400, detail="유효하지 않은 JSON 파일입니다.")

    # {"version":1, "config":{...}} 형식 또는 config 직접 포함 둘 다 허용
    config = data.get("config", data) if isinstance(data, dict) else None
    if not isinstance(config, dict):
        raise HTTPException(status_code=400, detail="설정 형식이 올바르지 않습니다.")

    value = json.dumps(config, ensure_ascii=False)
    setting = await crud.settings.get_setting(db, key=_SETTINGS_KEY)
    if setting is None:
        await crud.settings.create_setting(
            db,
            schemas.SettingsCreate(
                key=_SETTINGS_KEY,
                value=value,
                description='정책 삭제 워크플로우 설정 (fpat.yaml 형식)',
            ),
        )
    else:
        await crud.settings.update_setting(
            db=db,
            db_obj=setting,
            obj_in=schemas.SettingsUpdate(value=value),
        )

    _write_fpat_yaml(config)
    return {"ok": True, "message": "설정이 복구되었습니다."}


@router.post("/deletion-workflow/parse-yaml")
async def parse_yaml_to_json(request: Request):
    """YAML 텍스트를 JSON으로 파싱하여 반환합니다."""
    raw = await request.body()
    try:
        result = yaml.safe_load(raw.decode("utf-8"))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"YAML 파싱 오류: {e}")
    return {"data": result}


@router.get("/deletion-workflow/config/yaml")
async def get_deletion_workflow_config_yaml(db: AsyncSession = Depends(get_db)):
    """현재 삭제 워크플로우 설정을 YAML 텍스트로 반환합니다."""
    setting = await crud.settings.get_setting(db, key=_SETTINGS_KEY)
    if setting:
        try:
            config = _deep_merge(_default_config(), json.loads(setting.value))
        except Exception:
            config = _load_fpat_yaml()
    else:
        config = _load_fpat_yaml()

    yaml_text = yaml.dump(config, allow_unicode=True, default_flow_style=False, sort_keys=False)
    return Response(content=yaml_text, media_type="text/plain; charset=utf-8")


@router.put("/deletion-workflow/config/yaml")
async def update_deletion_workflow_config_yaml(
    request: Request,
    db: AsyncSession = Depends(get_db),
):
    """raw YAML 텍스트로 삭제 워크플로우 설정을 저장합니다."""
    raw = await request.body()
    try:
        yaml_text = raw.decode("utf-8")
        config = yaml.safe_load(yaml_text)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"YAML 파싱 오류: {e}")

    if not isinstance(config, dict):
        raise HTTPException(status_code=400, detail="YAML은 매핑(dict) 형식이어야 합니다.")

    value = json.dumps(config, ensure_ascii=False)
    setting = await crud.settings.get_setting(db, key=_SETTINGS_KEY)
    if setting is None:
        await crud.settings.create_setting(
            db,
            schemas.SettingsCreate(
                key=_SETTINGS_KEY,
                value=value,
                description='정책 삭제 워크플로우 설정 (fpat.yaml 형식)',
            ),
        )
    else:
        await crud.settings.update_setting(
            db=db,
            db_obj=setting,
            obj_in=schemas.SettingsUpdate(value=value),
        )

    _write_fpat_yaml(config)
    return {"ok": True}


# ──────────────────────────────────────────────
# 예외 설정 — 엑셀 일괄등록
# ──────────────────────────────────────────────

_EXCEPTION_CATEGORY_META = {
    "request_ids": {"key_field": "id", "key_label": "신청번호"},
    "static_list": {"key_field": "name", "key_label": "정책명"},
}


@router.get("/deletion-workflow/exceptions/{category}/excel-template")
async def download_exception_excel_template(category: str):
    """예외 설정(신청번호/정책명) 일괄등록용 엑셀 서식 파일 다운로드"""
    meta = _EXCEPTION_CATEGORY_META.get(category)
    if meta is None:
        raise HTTPException(status_code=404, detail="지원하지 않는 예외 카테고리입니다.")

    key_label = meta["key_label"]
    headers = [f"{key_label}*", "사유", "시작일 (YYYY-MM-DD)", "만료일 (YYYY-MM-DD)"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "예외 등록"

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    example_key = "REQ-1234" if category == "request_ids" else "allow_xxx"
    example_data = [
        [example_key, "임시예외", "", ""],
    ]
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    example_font = Font(size=10)
    for row_idx, row_data in enumerate(example_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.font = example_font

    ws2 = wb.create_sheet("설명")
    instructions = [
        ["필드명", "설명", "필수 여부", "예시"],
        [key_label, f"{key_label} 값", "필수", example_key],
        ["사유", "예외 등록 사유", "선택", "임시예외"],
        ["시작일", "예외 적용 시작일 (비우면 즉시 적용)", "선택", "2026-01-01"],
        ["만료일", "예외 적용 만료일 (비우면 무기한)", "선택", "2026-12-31"],
        ["", "", "", ""],
        ["주의사항", "", "", ""],
        ["- * 표시된 필드는 필수 입력 항목입니다.", "", "", ""],
        ["- 예시 행은 삭제하고 실제 데이터를 입력하세요.", "", "", ""],
    ]
    for row_idx, row_data in enumerate(instructions, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

    column_widths = [20, 30, 22, 22]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{category}_template.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/deletion-workflow/exceptions/{category}/excel-import")
async def import_exception_excel(category: str, file: UploadFile = File(...)):
    """엑셀 파일을 파싱해 예외 항목 리스트를 반환합니다 (DB에는 저장하지 않음 — 프론트에서 병합 후 별도 저장)."""
    meta = _EXCEPTION_CATEGORY_META.get(category)
    if meta is None:
        raise HTTPException(status_code=404, detail="지원하지 않는 예외 카테고리입니다.")
    if not file.filename or not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.")

    key_field = meta["key_field"]
    key_label = meta["key_label"]
    key_header = f"{key_label}*"

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active

        header_row = [cell.value for cell in ws[1]]
        header_mapping = {
            key_header: key_field,
            "사유": "reason",
            "시작일 (YYYY-MM-DD)": "start",
            "만료일 (YYYY-MM-DD)": "until",
        }

        if key_header not in header_row:
            raise HTTPException(status_code=400, detail=f"필수 컬럼이 없습니다: {key_header}")

        items = []
        errors = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):
                continue

            item = {}
            missing_key = False
            for col_idx, header in enumerate(header_row):
                if header not in header_mapping:
                    continue
                field_name = header_mapping[header]
                value = row[col_idx]

                if isinstance(value, str):
                    value = value.strip()
                    if not value:
                        value = None
                elif value is not None and hasattr(value, "strftime"):
                    value = value.strftime("%Y-%m-%d")
                elif field_name == key_field and isinstance(value, (int, float)):
                    # 신청번호/정책명을 숫자로만 입력한 경우 openpyxl이 int/float로 읽어오므로
                    # is_excepted()의 문자열 포함 비교(`id in value`)가 깨지지 않도록 문자열로 정규화
                    value = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)

                if header == key_header and not value:
                    errors.append(f"{row_idx}행: {key_header} 필드는 필수입니다.")
                    missing_key = True
                    break

                if value is not None:
                    item[field_name] = value

            if missing_key:
                continue
            if key_field not in item:
                continue
            item.setdefault("reason", "")
            items.append(item)

        if errors:
            raise HTTPException(status_code=400, detail="\n".join(errors))
        if not items:
            raise HTTPException(status_code=400, detail="등록할 예외 데이터가 없습니다.")

        return items
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"예외 엑셀 파싱 실패: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"엑셀 파싱 실패: {str(e)}")
