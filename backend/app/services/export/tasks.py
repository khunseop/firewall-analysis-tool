"""
Devices 페이지 "직접 추출"(단건/다건, 병합 포함)을 백그라운드에서 처리하는 오케스트레이터.

sync/tasks.py의 run_sync_all_orchestrator와 동일한 패턴(자체 SessionLocal 세션 + 단계별
상태 갱신 + WebSocket 브로드캐스트)을 따른다. 진행 상태는 ExportTask 테이블에 저장되고
websocket_manager를 통해 모든 클라이언트에 실시간 전파된다.
"""
import asyncio
import logging
from datetime import date, datetime
from io import BytesIO
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app import crud, models
from app.core.config import PROJECT_ROOT
from app.core.executors import IO_EXECUTOR
from app.db.session import SessionLocal
from app.services.sync.collector import create_collector_from_device
from app.services.websocket_manager import websocket_manager

logger = logging.getLogger(__name__)

EXPORT_TYPE_LABEL = {"policies": "정책", "objects": "객체", "hit_dates": "사용이력"}
EXPORT_DIR = PROJECT_ROOT / "exports"

_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

_POLICY_COL_MAP = {
    "vsys": "VSYS",
    "seq": "#",
    "rule_name": "정책명",
    "enable": "활성",
    "action": "액션",
    "source": "출발지",
    "destination": "목적지",
    "service": "서비스",
    "user": "사용자",
    "application": "애플리케이션",
    "security_profile": "보안 프로파일",
    "category": "카테고리",
    "description": "설명",
    "last_hit_date": "마지막 사용일",
}


def _now_kst() -> datetime:
    return datetime.now(ZoneInfo("Asia/Seoul")).replace(tzinfo=None)


def _write_df_to_ws(ws, df: pd.DataFrame) -> None:
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 20
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)


def _single_sheet_excel(df: pd.DataFrame, sheet_name: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_df_to_ws(ws, df)
    buf = BytesIO()
    wb.save(buf)
    return buf


def _multi_sheet_excel(sheets: dict[str, pd.DataFrame]) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)
        _write_df_to_ws(ws, df)
    buf = BytesIO()
    wb.save(buf)
    return buf


def _normalize_policy_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "enable" in df.columns:
        df["enable"] = df["enable"].map(lambda v: "활성" if str(v).upper() == "Y" else "비활성")
    return df.rename(columns={k: v for k, v in _POLICY_COL_MAP.items() if k in df.columns})


def _normalize_object_dfs(
    net_obj: pd.DataFrame,
    net_grp: pd.DataFrame,
    svc_obj: pd.DataFrame,
    svc_grp: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    return {
        "주소객체": net_obj.rename(columns={"Name": "이름", "Type": "타입", "Value": "IP 주소"}),
        "주소그룹": net_grp.rename(columns={"Group Name": "이름", "Entry": "멤버"}),
        "서비스객체": svc_obj.rename(columns={"Name": "이름", "Protocol": "프로토콜", "Port": "포트"}),
        "서비스그룹": svc_grp.rename(columns={"Group Name": "이름", "Entry": "멤버"}),
    }


async def _collect_db_policies(db, device_id: int) -> pd.DataFrame:
    policies = await crud.policy.get_policies_by_device(db, device_id=device_id)
    rows = [{
        "vsys": p.vsys, "seq": p.seq, "rule_name": p.rule_name,
        "enable": "Y" if p.enable else "N", "action": p.action,
        "source": p.source, "destination": p.destination, "service": p.service,
        "user": p.user, "application": p.application, "security_profile": p.security_profile,
        "category": p.category, "description": p.description, "last_hit_date": p.last_hit_date,
    } for p in policies]
    return pd.DataFrame(rows)


async def _collect_db_objects(db, device_id: int) -> dict[str, pd.DataFrame]:
    net_objs = await crud.network_object.get_all_active_network_objects_by_device(db, device_id=device_id)
    net_grps = await crud.network_group.get_all_active_network_groups_by_device(db, device_id=device_id)
    svc_objs = await crud.service.get_all_active_services_by_device(db, device_id=device_id)
    svc_grps = await crud.service_group.get_all_active_service_groups_by_device(db, device_id=device_id)

    net_obj_df = pd.DataFrame([{"Name": o.name, "Type": o.type, "Value": o.ip_address} for o in net_objs])
    net_grp_df = pd.DataFrame([{"Group Name": g.name, "Entry": g.members} for g in net_grps])
    svc_obj_df = pd.DataFrame([{"Name": s.name, "Protocol": s.protocol, "Port": s.port} for s in svc_objs])
    svc_grp_df = pd.DataFrame([{"Group Name": g.name, "Entry": g.members} for g in svc_grps])
    return _normalize_object_dfs(net_obj_df, net_grp_df, svc_obj_df, svc_grp_df)


async def _collect_db_hit_dates(db, device_id: int) -> pd.DataFrame:
    policies = await crud.policy.get_policies_by_device(db, device_id=device_id)
    rows = [{"vsys": p.vsys, "rule_name": p.rule_name, "last_hit_date": p.last_hit_date} for p in policies]
    return pd.DataFrame(rows)


async def _collect_hit_dates(
    collector,
    device: models.Device,
    use_ssh: bool,
    loop: asyncio.AbstractEventLoop,
    timeout: int,
) -> pd.DataFrame:
    def _method():
        if use_ssh:
            return collector.export_last_hit_date_ssh(timeout=timeout)
        return collector.export_last_hit_date()

    main_df: pd.DataFrame = await asyncio.wait_for(
        loop.run_in_executor(IO_EXECUTOR, _method), timeout=timeout
    )

    if not device.ha_peer_ip:
        return main_df

    ha_collector = create_collector_from_device(device, use_ha_ip=True)
    ha_df: pd.DataFrame | None = None
    try:
        await loop.run_in_executor(IO_EXECUTOR, ha_collector.connect)

        def _ha_method():
            if use_ssh:
                return ha_collector.export_last_hit_date_ssh(timeout=timeout)
            return ha_collector.export_last_hit_date()

        ha_df = await loop.run_in_executor(IO_EXECUTOR, _ha_method)
    except Exception as e:
        logger.warning(f"HA peer hit_date 수집 실패 ({device.ha_peer_ip}): {e}")
    finally:
        try:
            await loop.run_in_executor(IO_EXECUTOR, ha_collector.disconnect)
        except Exception:
            pass

    if ha_df is None or ha_df.empty:
        return main_df

    combined = pd.concat([main_df, ha_df])
    combined["last_hit_date"] = pd.to_datetime(combined["last_hit_date"], errors="coerce")
    subset = ["vsys", "rule_name"] if "vsys" in combined.columns else ["rule_name"]
    combined = combined.sort_values("last_hit_date", ascending=False, na_position="last")
    return combined.drop_duplicates(subset=subset, keep="first")


async def _collect_live_export(
    device: models.Device,
    export_type: str,
    use_ssh: bool,
    loop: asyncio.AbstractEventLoop,
    timeout: int,
) -> Any:
    try:
        collector = create_collector_from_device(device)
    except Exception as e:
        raise RuntimeError(f"Collector 초기화 실패 ({device.name}): {e}") from e

    try:
        await asyncio.wait_for(loop.run_in_executor(IO_EXECUTOR, collector.connect), timeout=min(timeout, 60))
    except asyncio.TimeoutError as e:
        raise RuntimeError(f"장비 연결 타임아웃 ({device.name})") from e
    except Exception as e:
        raise RuntimeError(f"장비 연결 실패 ({device.name}): {e}") from e

    try:
        if export_type == "policies":
            df = await asyncio.wait_for(
                loop.run_in_executor(IO_EXECUTOR, collector.export_security_rules), timeout=timeout
            )
            return _normalize_policy_df(df)
        elif export_type == "objects":
            net_obj = await loop.run_in_executor(IO_EXECUTOR, collector.export_network_objects)
            net_grp = await loop.run_in_executor(IO_EXECUTOR, collector.export_network_group_objects)
            svc_obj = await loop.run_in_executor(IO_EXECUTOR, collector.export_service_objects)
            svc_grp = await loop.run_in_executor(IO_EXECUTOR, collector.export_service_group_objects)
            return _normalize_object_dfs(net_obj, net_grp, svc_obj, svc_grp)
        else:
            return await _collect_hit_dates(collector, device, use_ssh, loop, timeout)
    finally:
        try:
            await loop.run_in_executor(IO_EXECUTOR, collector.disconnect)
        except Exception:
            pass


async def _update_export_task(
    task_id: int,
    *,
    status: str | None = None,
    step: str | None = None,
    progress_current: int | None = None,
    progress_total: int | None = None,
    error_message: str | None = None,
    result_file_path: str | None = None,
    result_filename: str | None = None,
) -> None:
    """ExportTask 상태를 갱신하고 WebSocket으로 브로드캐스트한다."""
    async with SessionLocal() as db:
        task = await db.get(models.ExportTask, task_id)
        if not task:
            return
        if status is not None:
            task.status = status
            if status == "in_progress" and task.started_at is None:
                task.started_at = _now_kst()
            if status in ("success", "failure"):
                task.completed_at = _now_kst()
        if step is not None:
            task.step = step
        if progress_current is not None:
            task.progress_current = progress_current
        if progress_total is not None:
            task.progress_total = progress_total
        if error_message is not None:
            task.error_message = error_message
        if result_file_path is not None:
            task.result_file_path = result_file_path
        if result_filename is not None:
            task.result_filename = result_filename

        db.add(task)
        await db.commit()
        await db.refresh(task)

        await websocket_manager.broadcast_export_status(
            task_id=task.id,
            status=task.status,
            step=task.step,
            progress_current=task.progress_current,
            progress_total=task.progress_total,
            error=task.error_message,
        )


async def run_export_task(task_id: int) -> None:
    """
    ExportTask 1건을 처리하는 오케스트레이터.

    1. 대상 장비 목록 로드
    2. 장비별 데이터 수집 (source='live'면 실시간 접속, 'db'면 동기화된 데이터 사용)
    3. 엑셀 생성 및 디스크 저장 (병합 옵션 처리)
    4. 최종 상태 반영 (성공/실패)
    """
    async with SessionLocal() as db:
        task = await db.get(models.ExportTask, task_id)
        if not task:
            logger.error(f"[export] ExportTask {task_id}를 찾을 수 없습니다.")
            return
        device_ids: list[int] = task.device_ids
        export_type: str = task.export_type
        source: str = task.source
        merge: bool = task.merge
        use_ssh: bool = task.use_ssh
        timeout: int = task.timeout_seconds

        devices = []
        for device_id in device_ids:
            device = await crud.device.get_device(db, device_id=device_id)
            if device:
                devices.append(device)

    if not devices:
        await _update_export_task(task_id, status="failure", error_message="대상 장비를 찾을 수 없습니다.")
        return

    await _update_export_task(
        task_id, status="in_progress", step="시작 중...", progress_current=0, progress_total=len(devices)
    )

    loop = asyncio.get_running_loop()
    per_device_data: dict[int, Any] = {}
    label = EXPORT_TYPE_LABEL[export_type]

    try:
        for idx, device in enumerate(devices, start=1):
            await _update_export_task(task_id, step=f"{device.name} 처리 중 ({idx}/{len(devices)})")
            if source == "db":
                async with SessionLocal() as db:
                    if export_type == "policies":
                        per_device_data[device.id] = _normalize_policy_df(await _collect_db_policies(db, device.id))
                    elif export_type == "objects":
                        per_device_data[device.id] = await _collect_db_objects(db, device.id)
                    else:
                        per_device_data[device.id] = await _collect_db_hit_dates(db, device.id)
            else:
                per_device_data[device.id] = await _collect_live_export(device, export_type, use_ssh, loop, timeout)
            await _update_export_task(task_id, progress_current=idx)

        await _update_export_task(task_id, step="엑셀 생성 중...")
        today = date.today().strftime("%Y%m%d")

        if merge and len(devices) > 1:
            merged_sheets: dict[str, pd.DataFrame] = {}
            for device in devices:
                data = per_device_data[device.id]
                if isinstance(data, dict):
                    for sheet_name, df in data.items():
                        merged_sheets[f"{sheet_name}_{device.name}"[:31]] = df
                else:
                    merged_sheets[device.name[:31]] = data
            output = _multi_sheet_excel(merged_sheets)
            filename = f"통합_{label}_{today}.xlsx"
        else:
            device = devices[0]
            data = per_device_data[device.id]
            output = _multi_sheet_excel(data) if isinstance(data, dict) else _single_sheet_excel(data, label)
            filename = f"{device.name}_{label}_{today}.xlsx"

        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        file_path = EXPORT_DIR / f"{task_id}.xlsx"
        output.seek(0)
        file_path.write_bytes(output.read())

        await _update_export_task(
            task_id,
            status="success",
            step="완료",
            progress_current=len(devices),
            result_file_path=str(file_path),
            result_filename=filename,
        )
    except asyncio.TimeoutError:
        await _update_export_task(task_id, status="failure", error_message=f"데이터 수집 타임아웃 ({timeout}초 초과)")
    except NotImplementedError:
        await _update_export_task(task_id, status="failure", error_message="이 장비는 해당 기능을 지원하지 않습니다.")
    except Exception as e:
        logger.error(f"[export] 작업 실패 task_id={task_id}: {e}", exc_info=True)
        await _update_export_task(task_id, status="failure", error_message=str(e))
