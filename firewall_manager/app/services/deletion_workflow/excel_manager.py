"""
Excel 파일 관리 유틸리티
"""
import logging
from pathlib import Path
from typing import Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Font

logger = logging.getLogger(__name__)


class ExcelManager:
    """Excel 파일 관리 클래스"""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        ExcelManager 초기화
        
        Args:
            config: 설정 딕셔너리 (None이면 기본값 사용)
        """
        self.config = config or {}
    
    def _get_config_value(self, key: str, default: Any = None) -> Any:
        """
        설정값 가져오기 (점으로 구분된 경로 지원)
        
        Args:
            key: 설정 키 (예: 'excel_styles.header_fill_color')
            default: 기본값
            
        Returns:
            설정값 또는 기본값
        """
        keys = key.split('.')
        value = self.config
        
        try:
            for k in keys:
                value = value[k]
            return value
        except (KeyError, TypeError):
            logger.debug(f"설정 키 '{key}'를 찾을 수 없습니다. 기본값 '{default}'를 사용합니다.")
            return default
    
    def save_dataframe_to_excel(
        self,
        df: pd.DataFrame,
        file_path: str,
        sheet_name: str = "Sheet1",
        index: bool = False,
        style: bool = True
    ) -> bool:
        """
        DataFrame을 Excel 파일로 저장
        
        Args:
            df: 저장할 DataFrame
            file_path: 저장할 파일 경로
            sheet_name: 시트 이름
            index: 인덱스 포함 여부
            style: 스타일 적용 여부
            
        Returns:
            저장 성공 여부
        """
        try:
            # 파일 경로의 디렉토리가 없으면 생성
            file_path_obj = Path(file_path)
            file_path_obj.parent.mkdir(parents=True, exist_ok=True)
            
            # Excel 파일로 저장
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=index)
            
            # 스타일 적용
            if style:
                self._apply_excel_styles(file_path, sheet_name)
            
            logger.info(f"Excel 파일 저장 완료: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Excel 파일 저장 실패: {file_path}, 오류: {e}")
            return False
    
    def _apply_excel_styles(self, file_path: str, sheet_name: str) -> None:
        """
        Excel 파일에 스타일 적용
        
        Args:
            file_path: Excel 파일 경로
            sheet_name: 시트 이름
        """
        try:
            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                logger.warning(f"시트 '{sheet_name}'를 찾을 수 없습니다.")
                return
            
            sheet = wb[sheet_name]
            
            # 첫 번째 행에 통계 정보 추가 (있는 경우)
            if sheet.max_row > 0:
                # 헤더 스타일 설정
                header_color = self._get_config_value('excel_styles.header_fill_color', 'E0E0E0')
                history_color = self._get_config_value('excel_styles.history_fill_color', 'ccffff')
                
                # 헤더 행 스타일 적용 (첫 번째 데이터 행)
                header_row = 1
                if sheet.max_row > 0:
                    for col in range(1, min(8, sheet.max_column + 1)):
                        cell = sheet.cell(row=header_row, column=col)
                        cell.alignment = Alignment(horizontal='center')
                        cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
                        cell.font = Font(bold=True)
                    
                    # 이력 컬럼 스타일 적용 (8번째 컬럼부터)
                    if sheet_name != '이력없음_미사용정책':
                        for col in range(8, min(24, sheet.max_column + 1)):
                            cell = sheet.cell(row=header_row, column=col)
                            cell.fill = PatternFill(start_color=history_color, end_color=history_color, fill_type='solid')
            
            wb.save(file_path)
            logger.debug(f"Excel 스타일 적용 완료: {file_path}")
        except Exception as e:
            logger.warning(f"Excel 스타일 적용 실패: {file_path}, 오류: {e}")
    
    def read_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        **kwargs
    ) -> pd.DataFrame:
        """
        Excel 파일 읽기
        
        Args:
            file_path: Excel 파일 경로
            sheet_name: 시트 이름 (None이면 첫 번째 시트)
            **kwargs: pandas.read_excel에 전달할 추가 인자
            
        Returns:
            읽은 DataFrame
        """
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name, **kwargs)
            else:
                df = pd.read_excel(file_path, **kwargs)
            logger.debug(f"Excel 파일 읽기 완료: {file_path}")
            return df
        except Exception as e:
            logger.error(f"Excel 파일 읽기 실패: {file_path}, 오류: {e}")
            raise
    
    def read_excel_all_sheets(self, file_path: str) -> Dict[str, pd.DataFrame]:
        """
        Excel 파일의 모든 시트 읽기
        
        Args:
            file_path: Excel 파일 경로
            
        Returns:
            시트 이름을 키로 하는 DataFrame 딕셔너리
        """
        try:
            excel_file = pd.ExcelFile(file_path)
            sheets = {}
            for sheet_name in excel_file.sheet_names:
                sheets[sheet_name] = pd.read_excel(excel_file, sheet_name=sheet_name)
            logger.debug(f"Excel 파일의 모든 시트 읽기 완료: {file_path}, 시트 수: {len(sheets)}")
            return sheets
        except Exception as e:
            logger.error(f"Excel 파일 읽기 실패: {file_path}, 오류: {e}")
            raise
    
    def save_to_excel_with_sheets(
        self,
        data_dict: Dict[str, pd.DataFrame],
        file_path: str,
        style: bool = True
    ) -> bool:
        """
        여러 DataFrame을 각각 다른 시트로 저장
        
        Args:
            data_dict: 시트 이름을 키로 하는 DataFrame 딕셔너리
            file_path: 저장할 파일 경로
            style: 스타일 적용 여부
            
        Returns:
            저장 성공 여부
        """
        try:
            # 파일 경로의 디렉토리가 없으면 생성
            file_path_obj = Path(file_path)
            file_path_obj.parent.mkdir(parents=True, exist_ok=True)
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 각 시트에 스타일 적용
            if style:
                for sheet_name in data_dict.keys():
                    self._apply_excel_styles(file_path, sheet_name)
            
            logger.info(f"Excel 파일 저장 완료 (시트 수: {len(data_dict)}): {file_path}")
            return True
        except Exception as e:
            logger.error(f"Excel 파일 저장 실패: {file_path}, 오류: {e}")
            return False
    
    def add_summary_row(
        self,
        file_path: str,
        sheet_name: str,
        summary_text: str,
        column: str = "A"
    ) -> bool:
        """
        Excel 파일의 첫 번째 행에 요약 정보 추가
        
        Args:
            file_path: Excel 파일 경로
            sheet_name: 시트 이름
            summary_text: 요약 텍스트
            column: 요약 정보를 추가할 컬럼 (기본값: "A")
            
        Returns:
            추가 성공 여부
        """
        try:
            wb = load_workbook(file_path)
            if sheet_name not in wb.sheetnames:
                logger.warning(f"시트 '{sheet_name}'를 찾을 수 없습니다.")
                return False
            
            sheet = wb[sheet_name]
            
            # 첫 번째 행 삽입
            sheet.insert_rows(1)
            
            # 요약 정보 추가
            summary_cell = sheet[f"{column}1"]
            summary_cell.value = summary_text
            summary_cell.font = Font(bold=True)
            
            wb.save(file_path)
            logger.debug(f"요약 행 추가 완료: {file_path}")
            return True
        except Exception as e:
            logger.warning(f"요약 행 추가 실패: {file_path}, 오류: {e}")
            return False

